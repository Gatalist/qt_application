from openpyxl import load_workbook
import pandas
import re
from settings import Settings
import json
import codecs
import os
import threading


class JsonDocument(Settings):
    _instance = None
    def __new__(class_, *args, **kwargs):
        if not isinstance(class_._instance, class_):
            class_._instance = object.__new__(class_, *args, **kwargs)
        return class_._instance
    
    def __init__(self) -> None:
        self.format_open = 'Json (*.json)'
        self.document: str = None # ссылка на документ json
        self.all_list_admin_value = []
    
    def load_data_from_file(self, file_name):
        self.document = file_name
        threading.Thread(target = self.get_values_from_json).start()

    def get_values_from_json(self):
        file_ = json.load(codecs.open(filename=self.document, mode='r', encoding='utf-8'))
        name_values = []
        for value in file_["data"]:
            name_values.append(value['name'])
        self.all_list_admin_value = name_values


class ExcelDocument(Settings):
    _instance = None
    def __new__(class_, *args, **kwargs):
        if not isinstance(class_._instance, class_):
            class_._instance = object.__new__(class_, *args, **kwargs)
        return class_._instance
    
    def __init__(self) -> None:
        self.format_open = 'Excel (*.xlsx);;Excel (*.xls)'
        self.document: str = None # ссылка на документ xlsx
        
        self.data_frame: object = None  # получаем data_frame документа в pandas
        self.work_book: object = None  # открываем документ в openpyxl
        self.list_sheet: list = []  # получаем список листов в документе
        self.work_sheet: object = None  # получаем рабочий лист в документе
        self.count_row: int = None  # получаем список строк в документе
        self.start_row: int = 2 # с какой строки начинать читать документ
        self.list_row: list = [] # список колонок на листе
        self.list_column: list = [] # список колонок на листе

    def load_data_from_file(self, file_name):
        self.document = file_name
        self.work_book = load_workbook(self.document)  # открываем документ в openpyxl
        self.list_sheet = self.work_book.sheetnames # получаем список листов в документе
        self.work_sheet = self.work_book[self.list_sheet[0]]  # делаем 1 лист активным по умолчанию
        self.load_data_from_sheet(sheet_name=self.list_sheet[0])

    def load_data_from_sheet(self, sheet_name):
        self.data_frame = pandas.read_excel(self.document, sheet_name=sheet_name) # получаем data_frame документа в pandas
        self.count_row = len(self.data_frame.index) + 1 # получаем список строк в листе
        self.list_row = [string for string in range(self.start_row, self.count_row + 1)] # генерируем список строк [начало, конец]
        self.list_column = self.data_frame.columns.values.tolist() # получаем все колонки на листе

    def active_list_to_write(self, sheet_name):
        self.work_sheet = self.work_book[sheet_name]  # делаем лист активным по умолчанию
        self.load_data_from_sheet(sheet_name=sheet_name)

    # возвращаем список нумерованых строк с данными в указаной ячейке
    def get_rows_from_column(self, column_name) -> list:
        data_rows = []
        start_number_string = self.start_row
        for row in self.data_frame[column_name]:
            row_to_line = []
            if type(row) == str:
                row_to_line.append(start_number_string)
                row_to_line.append(row)
                data_rows.append(row_to_line)
            start_number_string += 1
        return data_rows

    # получаем обьект ячейки по координатам: например 'AC4'
    def get_cell_obj(self, cell_letter, cell_number) -> object:
        cell = f'{cell_letter}{cell_number}'
        return self.work_sheet[cell]
    
    # сохраняем результат в ячейку
    def save_result_in_cell(self, cell_object: object, text: str) -> object:
        self.work_sheet[cell_object.coordinate] = text
        return self.work_sheet


class ReadExcelDocument:
    symbol_arrow = '👉'
    symbol_ok = '✅'
    symbol_warning = '⚠️'
    symbol_forbidden = '⛔️'

    @staticmethod
    # возвращаем список нумерованых строк с данными в указаной ячейке
    def out_text(number_string, message) -> str:
        return f'-----[ Строка: {number_string} ]-----\n{message}\n'
    
    # Выводим данные ячейки. Аргумент "read_line" разобъет текст на строки по символу ";"
    def read_list_data_row(self, list_data: str) -> str:
        number = 0
        if len(list_data) > 0:
            for number_string, text in list_data:              
                string = ''
                call_data = text.split(';')
                for line in call_data:
                    string += f'{line}\n'
                number += 1
                yield self.out_text(number_string, string)
        else:
            yield f"{self.symbol_forbidden} Колонка пустая\n"
        yield f"{self.symbol_ok} Прочитано строк {self.symbol_arrow} {number}\n\n"
    
    # проверяем строку на первую заглавную букву
    def check_error_in_column(self, list_data) -> str:
        errors = 0
        for number_string, text in list_data:            
            string = ''
            check = False
            if text:
                for line in text.split(';'):
                    if len(line) < 1:
                        check = True
                        string += f'{self.symbol_warning};\n'
                        errors += 1
                    else:
                        if not line[0].istitle() and not line[0].isdigit():
                            check = True
                            string += f'{self.symbol_warning}{line}\n'
                            errors += 1
                        else:
                            string += f'{line}\n'
                if check:
                    yield self.out_text(number_string, string)
        if errors:
            yield f"❌ Ошибок {self.symbol_arrow} {errors}\n\n"
        else:
            yield f"{self.symbol_ok} Ошибок не обнаружено\n\n"

    # ищим фрагмент текста в ячейке
    def search_text(self, list_data, search: str):
        result = 0
        for number_string, text in list_data:
            like = False
            string = ''
            for element in text.split(';'):
                if element.lower().find(search.lower()) != -1:
                    string += f"{element}\n"
                    result += 1
                    like = True
            if like:
                yield self.out_text(number_string, string)
        yield f"{self.symbol_ok} Найдено совпадений {self.symbol_arrow} {result}"

    # получаем уникальные строки ячейки
    def get_unique_strings(self, list_data):
        unique_elem = []
        for number_string, text in list_data:
            for line in text.split(';'):
                if line not in unique_elem:
                    unique_elem.append(line)
        
        for elem in unique_elem:
            if elem == "":
                yield ";\n"
            else:
                yield f"{elem}\n"
        yield f"\n{self.symbol_ok} Уникальных строк {self.symbol_arrow} {len(unique_elem)}"

    def get_unused_value_in_admin(self, list_data_xl, list_data_admin):
        unique_elem = []
        for number_string, text in list_data_xl:
            for line in text.split(';'):
                if line not in unique_elem:
                    unique_elem.append(line)

        unused_value_for_admin = list(filter(lambda x: x not in unique_elem, list_data_admin))
        
        for elem in unused_value_for_admin:
            if elem == "":
                yield ";\n"
            else:
                yield f"{elem}\n"
        
        yield f"\n✅ Не используемых значений -> {len(unused_value_for_admin)}"


class WriteExcelDocument:
    symbols_1 = [";;", "; ;", ";  ;" ";;;", ";;;;"]
    symbols_2 = ["  ", "   ", "    ", "\t"] #, "\r\n", "\n", "\r"]

    # делаем первую букву каждой строки заглавной
    def upper_first_letter_in_text(self, text: str) -> str:
        split_text = text.split(';')
        new_list = []
        for line in split_text:
            capitalized = line[0:1].upper() + line[1:]
            new_list.append(capitalized)
        return ';'.join(new_list)
    
    # заменяем сымволы в строке
    def replace_symbol(self, text: str) -> str:
        for symbol in self.symbols_1:
            text.replace(symbol, ';')

        for symbol in self.symbols_2:
            text.replace(symbol, ' ')

        text.strip()

        if len(text) > 0:
            if text[0] == ';':
                text = text[1:]
        if len(text) > 0:
            if text[-1] == ';':
                text = text[:-1]
        return text.strip()
    
    def out_text(self, number_string, message):
        return f'-----[ Строка: {number_string} ]-----\n{message}\n'
        # return f'<span style="color: #228B22;">{text}</span>'
    
    # меняем текст ячейки на новый
    # def change_text_to_cell(self, document, cell_object: object, text: str) -> object:
    #         document.save_result_in_cell(cell_object, upper_first_letter)

    #         document.save_result_in_cell(cell_object, None)
    #     return document

    def delete_symbol_enter(self, text):
        pattern = r'[;\n]'
        # Разделение строки по указанным разделителям
        result = re.split(pattern, text)
        # Фильтрация пустых строк из результата
        result = list(filter(None, result))
        return ";".join(result)

    # добавление фрагмента текста
    def add_text(self, document, cell_object: object, text: str) -> object:
        if text:
            cell_object_value = cell_object.value
            if cell_object_value is not None:
                cell_object_value = f'{cell_object_value};{text}'
            else:
                cell_object_value = text

            clear = self.delete_symbol_enter(cell_object_value)
            clear_txt = self.replace_symbol(clear)
            upper_first_letter = self.upper_first_letter_in_text(clear_txt)
            document.save_result_in_cell(cell_object, upper_first_letter)
        else:
            document.save_result_in_cell(cell_object, None)
        return document

    # добавление фрагмента текста в конец в ячейки
    def add_text_from_position(self, document: object, cell_object: object, text: str, position: str) -> object:
        if text:
            cell_object_value = cell_object.value
            if position == 'start' and cell_object_value is not None:
                cell_object_value = f'{text};{cell_object_value}'
               
            if position == 'end' and cell_object_value is not None:
                cell_object_value = f'{cell_object_value};{text}'
              
            if position == 'all' and cell_object_value is not None:
                cell_object_value = f'{text};{cell_object_value}'
            else:
                cell_object_value = text
                
            document.save_result_in_cell(cell_object, cell_object_value)
        else:
            document.save_result_in_cell(cell_object, None)
        return document

    ######################

    # вырезаем весь текст с одной ячейки и добавляем в другую
    def move_text_to_another_cell(self, document, cell_move: str, cell_past: str):
        result = 0
        for number_string in document.list_row:
            cell_move_obj = document.get_cell_obj(cell_move, number_string)
            print(cell_move_obj.value)
            # вырезаем данные с ячейки если она не пустая
            if cell_move_obj.value is not None:
                save_to_return_text = cell_move_obj.value
     
                # вставляем данные в другую ячейку
                cell_past_obj = document.get_cell_obj(cell_past, number_string)
                self.add_text(document, cell_past_obj, text=cell_move_obj.value)

                # очищаем ячейку откуда копируем текст
                document.save_result_in_cell(cell_move_obj, None)
                
                yield self.out_text(number_string, save_to_return_text)
                result += 1

        yield f"✅ Текст перемещен с ячеек [ {cell_move} ] в [ {cell_past} ] - {result}\n"

    # добавление фрагмента текста в ячейку
    def add_text_to_cell(self, document: object, cell_past: str, text: str, position: str):
        result = 0
        for number_string in document.list_row:
            cell_move_obj = document.get_cell_obj(cell_past, number_string)
            self.add_text_from_position(document, cell_move_obj, text, position)

            yield self.out_text(number_string, text)
            result += 1

        yield f"✅ Текст добавлен в колонку [ {cell_past} ] - {result} \n"

    # удаяем текст поиска с ячейки и добавляем в другую ячейку
    def move_search_text_to_other_cell(self, document: object, cell_move: str, cell_past: str, search: list):
        result = 0
        for number_string in document.list_row:
            cell_move_obj = document.get_cell_obj(cell_move, number_string)
            
            call_curent_text = cell_move_obj.value
            if call_curent_text is not None:
                list_search_text_lower = [word.lower() for word in search.split(';')]

                new_cell_list_text = []
                current_cell_list_text = []
                
                # перебираем список совпадений
                list_cur_text = call_curent_text.split(';')
                for search_word in list_search_text_lower:
                    # перебераем список строк
                    for line in list_cur_text:
                        if line.lower().find(search_word) != -1:
                            if line not in new_cell_list_text:
                                new_cell_list_text.append(line)
                        else:
                            current_cell_list_text.append(line)

                # удаляем найденные атрибуты с текущего текста
                new_curent_text = ";".join(current_cell_list_text)
                
                # save text in current cell
                document.save_result_in_cell(cell_move_obj, new_curent_text)

                # добавление фрагмента текста в другую ячейку
                cell_past_txt_add = ";".join(new_cell_list_text)
                cell_past_obj = document.get_cell_obj(cell_past, number_string)

                document.save_result_in_cell(cell_past_obj, cell_past_txt_add)

                yield self.out_text(number_string, new_curent_text)
                result += 1

        yield f"✅ Изменено строк - {result} \n"

