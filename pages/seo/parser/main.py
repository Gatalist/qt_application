import json
import requests
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill



class ProcessingData:
    def request_data(self, url: str) -> json:
        try:
            result = requests.get(url)
            if result.status_code == 200:
                return result.json()
            return {'error': result}
        except Exception as error:
            print(error)



# получение данных с сервера по ссылке (и API-key если нужен)
class CitrusApi(ProcessingData):

    # Запись данных в файл JSON
    def save_json(self, name, data):
        with open(f'{name}.json', 'w', encoding='utf-8') as output_file:
            json.dump(data, output_file, indent=4)  # indent=4 для удобного читаемого формата

    def parse_category_card(self, category_url, page_ctart, page_end, name_save=None):
        card_list = []
        for page in range(page_ctart, page_end + 1):
            url = f'https://api.ctrs.com.ua/router?with_meta=1&l=uk&url=/{category_url}/page_{page}/'
            print(url)

            page_data = self.request_data(url)

            data = page_data["data"]
            facetObject = data["facetObject"]
            items = facetObject["items"]

            card_list.extend(items)
        
        if name_save:
            self.save_json(name_save, card_list)
        return card_list

    def get_data_card(self, data, content_type, name_save):
        card_list_data = []
        for item in data:
            if item['preview']['src'].endswith(content_type):
                card_list_data.append({
                    "id": item['id'],
                    "url": item['url'],
                    "src": item['preview']['src']
                })

        if name_save:
            self.save_json(name_save + content_type, card_list_data)





class SpeedPageTest:
    def __init__(self, sheet_name):
        self.base_url = 'https://www.ctrs.com.ua'  # citrus base url page
        self.api_key = 'AIzaSyBEqZT1ymNblPkgSafDsPmQr9CMhKpsTf4'  # PageSpeed API key
        self.this_folder = os.getcwd()  # получаем текущую директорию

        self.work_book = Workbook()
        self.sheet = self.work_book.active
        self.sheet.title = sheet_name

        self.num = 2

        # устанавливаем название колонок
        self.sheet["A1"] = "id"
        self.sheet["B1"] = "url"
        self.sheet["C1"] = "first_contentful_paint"
        self.sheet["D1"] = "largest_contentful_paint"
        self.sheet["E1"] = "total_blocking_time"
        self.sheet["F1"] = "cumulative_layout_shift"
        self.sheet["G1"] = "speed_index"
        self.sheet["H1"] = "performance"

        # Установка ширины колонки
        # self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 90
        self.sheet.column_dimensions['C'].width = 25
        self.sheet.column_dimensions['D'].width = 25
        self.sheet.column_dimensions['E'].width = 25
        self.sheet.column_dimensions['F'].width = 25
        self.sheet.column_dimensions['G'].width = 15
        self.sheet.column_dimensions['H'].width = 15

        # цвета заливки ячеек
        self.color_bad = 'ff3333'
        self.color_normal = 'ffaa33'
        self.color_good = '00cc66'

        self.sum_value_first_contentful_paint = 0
        self.sum_value_largest_contentful_paint = 0
        self.sum_value_total_blocking_time = 0
        self.sum_value_cumulative_layout_shift = 0
        self.sum_value_speed_index = 0
        self.sum_value_performance = 0

    # получение данных с сервера по ссылке (и API-key если нужен)
    def request_data(self, url: str) -> json:
        try:
            result = requests.get(url)
            if result.status_code == 200:
                return result.json()
            return {}
        except Exception as error:
            print(error)

    # извлекаем данные с json ответа
    def select_data(self, card_id, result_speed_test):
        if result_speed_test:
            result = {}
            metrics = result_speed_test['lighthouseResult']['audits']
            
            result["id"] = card_id
            result["url"] = result_speed_test['id']
            result["first_contentful_paint"] = metrics['first-contentful-paint']['displayValue'].replace('\u00a0', ' ')
            result["largest_contentful_paint"] = metrics['largest-contentful-paint']['displayValue'].replace('\u00a0', ' ')
            result["total_blocking_time"] = metrics['total-blocking-time']['displayValue'].replace('\u00a0', ' ')
            result["cumulative_layout_shift"] = metrics['cumulative-layout-shift']['displayValue'].replace('\u00a0', ' ')
            result["speed_index"] = metrics['speed-index']['displayValue'].replace('\u00a0', ' ')
            result["performance"] = result_speed_test['lighthouseResult']['categories']['performance']['score']
            return result
        print('Bad request, not result')


    def worker(self, file_json_cards, mobile=False):
        # Указать путь к вашему файлу JSON
        file_path = os.path.join(self.this_folder, 'ctrs', file_json_cards)
        # Открыть файл JSON и загрузить его содержимое
        with open(file_path, 'r') as read_file:
            data = json.load(read_file)

        result = []
        for page in data:
            url = page['url']
            idd = page['id']
            url_to_analyze = self.base_url + url
            # endpoint = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url_to_analyze}&key={self.api_key}' # desctop
            endpoint = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url_to_analyze}&key={self.api_key}&strategy=mobile' # mobile
            # print(url_to_analyze)
            result_speed_test = self.request_data(endpoint)
            select_data = self.select_data(idd, result_speed_test)
            print(select_data)
            result.append(select_data)
        
        self.save_json('SpeedPageTest', result)
        print("save result SpeedPageTest")
        # return result

    # Создание объекта для заполнения (цвет фона)
    def color_fill_call(self, color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')  # Например, желтый цвет

    # проверяем значение и возвращаем цвет ячейки
    def first_contentful_paint_color(self, value):
        # Отлично: Менее 1 секунды | Хорошо: 1-2 секунды | Плохо: Более 2 секунд
        if value <= 1:
            return self.color_fill_call(self.color_good)
        if value > 1 and value < 2:
            return self.color_fill_call(self.color_normal)
        if value >= 2:
            return self.color_fill_call(self.color_bad)

    # проверяем значение и возвращаем цвет ячейки
    def largest_contentful_paint_color(self, value):
        # Отлично: Менее 2 секунд | Хорошо: 2-4 секунды | Плохо: Более 4 секунд
        if value <= 2:
            return self.color_fill_call(self.color_good)
        if value > 2 and value < 4:
            return self.color_fill_call(self.color_normal)
        if value >= 4:
            return self.color_fill_call(self.color_bad)

    # проверяем значение и возвращаем цвет ячейки
    def total_blocking_time_color(self, value):
        # Отлично: Менее 100 миллисекунд | Хорошо: 100-300 миллисекунд | Плохо: Более 300 миллисекунд
        if value <= 100:
            return self.color_fill_call(self.color_good)
        if value > 100 and value <= 300:
            return self.color_fill_call(self.color_normal)
        if value > 300:
            return self.color_fill_call(self.color_bad)

    # проверяем значение и возвращаем цвет ячейки
    def cumulative_layout_shift_color(self, value):
        # Отлично: Менее 0.1 | Хорошо: 0.1-0.25 | Плохо: Более 0.25
        if value <= 0.1:
            return self.color_fill_call(self.color_good)
        if value > 0.1 and value <= 0.25:
            return self.color_fill_call(self.color_normal)
        if value > 0.25:
            return self.color_fill_call(self.color_bad)

    # проверяем значение и возвращаем цвет ячейки
    def speed_index_color(self, value):
        # Отлично: Менее 1,000 | Хорошо: 1,000-1,500 | Плохо: Более 1,500
        if value <= 1.000:
            return self.color_fill_call(self.color_good)
        if value > 1.000 and value <= 1.500:
            return self.color_fill_call(self.color_normal)
        if value > 1.500:
            return self.color_fill_call(self.color_bad)

    # проверяем значение и возвращаем цвет ячейки
    def performance_color(self, value):
        # Отлично: 90 и выше | Хорошо: 50-89 | Плохо: Менее 50
        if value >= 0.9:
            return self.color_fill_call(self.color_good)
        if value >= 0.5 and value <= 0.89:
            return self.color_fill_call(self.color_normal)
        if value < 0.5:
            return self.color_fill_call(self.color_bad)

    def get_number_to_str(self, value):
        replace_s = value.replace(' s', '')
        replace_ms = replace_s.replace(' ms', '')
        replace_symbol = replace_ms.replace(',', '.')
        return float(replace_symbol)

    def complete_table(self, file_name):
   
        with open(f'{file_name}.json', 'r') as json_file:
            file_json = json.load(json_file)

        for card_info in file_json:
            self.sheet[f"A{self.num}"] = card_info["id"]
            self.sheet[f"B{self.num}"] = card_info["url"]
            
            c_val = card_info["first_contentful_paint"]
            c_number = self.get_number_to_str(c_val)
            self.sheet[f"C{self.num}"] = c_val
            self.sheet[f"C{self.num}"].fill = self.first_contentful_paint_color(c_number)
            self.sum_value_first_contentful_paint += c_number

            d_val = card_info["largest_contentful_paint"]
            d_number = self.get_number_to_str(d_val)
            self.sheet[f"D{self.num}"] = d_val
            self.sheet[f"D{self.num}"].fill = self.largest_contentful_paint_color(d_number)
            self.sum_value_largest_contentful_paint += d_number
            
            e_val = card_info["total_blocking_time"]
            e_number = e_val.replace(',', '').replace(' ms', '')
            self.sheet[f"E{self.num}"] = str(e_number) + ' ms'
            self.sheet[f"E{self.num}"].fill = self.total_blocking_time_color(int(e_number))
            self.sum_value_total_blocking_time += int(e_number)
            
            f_val = card_info["cumulative_layout_shift"]
            f_number = self.get_number_to_str(f_val)
            self.sheet[f"F{self.num}"] = f_val
            self.sheet[f"F{self.num}"].fill = self.cumulative_layout_shift_color(f_number)
            self.sum_value_cumulative_layout_shift += f_number
            
            g_val = card_info["speed_index"]
            g_number = self.get_number_to_str(g_val)
            self.sheet[f"G{self.num}"] = g_val
            self.sheet[f"G{self.num}"].fill = self.speed_index_color(g_number)
            self.sum_value_speed_index += g_number

            h_val = card_info["performance"]
            h_number = self.get_number_to_str(str(h_val))
            self.sheet[f"H{self.num}"] = str(h_number) + ' '
            self.sheet[f"H{self.num}"].fill = self.performance_color(h_number)
            self.sum_value_performance += h_number

            self.num += 1

    # получаем средние значение и вносим в таблицу последнуй строкой
    def complete_average_value(self):
        # получаем средние значение
        element_count = self.num - 2
        number_sum = self.num + 2

        average_first_contentful_paint = self.sum_value_first_contentful_paint / element_count
        self.sheet[f"C{number_sum}"] = str(round(average_first_contentful_paint, 1)).replace(',', '.') + ' s'
        self.sheet[f"C{number_sum}"].fill = self.first_contentful_paint_color(average_first_contentful_paint)

        average_largest_contentful_paint = self.sum_value_largest_contentful_paint / element_count
        self.sheet[f"D{number_sum}"] = str(round(average_largest_contentful_paint, 1)).replace(',', '.') + ' s'
        self.sheet[f"D{number_sum}"].fill = self.largest_contentful_paint_color(average_largest_contentful_paint)

        average_total_blocking_time = self.sum_value_total_blocking_time / element_count
        self.sheet[f"E{number_sum}"] = str(int(average_total_blocking_time)).replace(',', '.') + ' ms'
        self.sheet[f"E{number_sum}"].fill = self.total_blocking_time_color(average_total_blocking_time)

        average_cumulative_layout_shift = self.sum_value_cumulative_layout_shift / element_count
        self.sheet[f"F{number_sum}"] = str(round(average_cumulative_layout_shift, 3)).replace(',', '.')
        self.sheet[f"F{number_sum}"].fill = self.cumulative_layout_shift_color(average_cumulative_layout_shift)

        average_speed_index = self.sum_value_speed_index / element_count
        self.sheet[f"G{number_sum}"] = str(round(average_speed_index, 1)).replace(',', '.') + ' s'
        self.sheet[f"G{number_sum}"].fill = self.speed_index_color(average_speed_index)
        
        average_performance = self.sum_value_performance / element_count
        self.sheet[f"H{number_sum}"] = str(round(average_performance, 2)).replace(',', '.')
        self.sheet[f"H{number_sum}"].fill = self.performance_color(average_performance)

    # Запись данных в файл xlsx
    def save_xlsx(self, save_name):
        self.work_book.save(f'{save_name}.xlsx')
        print('save file')
    
    # Запись данных в файл JSON
    def save_json(self, name, data):
        with open(f'{name}.json', 'w', encoding='utf-8') as output_file:
            json.dump(data, output_file, indent=4)  # indent=4 для удобного читаемого формата



speed_page_test = SpeedPageTest(sheet_name="content jpg")
# data_json = speed_page_test.worker('noutbuki_jpg.json', mobile=True)  # получаем результаты с API SpeedPageTest
speed_page_test.complete_table(file_name='SpeedPageTest')  # json file
speed_page_test.complete_average_value() # вносим средние значение в таблицу последнуй строкой
speed_page_test.save_xlsx('noutbuki_jpg')  # save xlsx format







ctrs = CitrusApi()

name_save = 'noytbooki'
data = ctrs.parse_category_card(category_url='noutbuki-i-ultrabuki', page_ctart=1, page_end=2)  # , name_save=name_save)
ctrs.get_data_card(data=data, content_type='.webp', name_save=name_save)

