from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
import json, os
from .parser import Config



class Document(Config):
    def __init__(self, open_file):
        self.open_file = self.get_file(open_file)
        self.work_book = self.open_or_create_xl(self.open_file)
        self.sheet = self.work_book.active 

    def get_file(self, name):
        file_name = f'{name}.xlsx'
        return os.path.join(self.path_save, f"{file_name}")

    # добавляем первую строку, заголовки колонок
    def add_title_column(self, sheet_obj):
        # устанавливаем название колонок
        sheet_obj["A1"] = "id"
        sheet_obj["B1"] = "url"
        sheet_obj["C1"] = "first_contentful_paint"
        sheet_obj["D1"] = "largest_contentful_paint"
        sheet_obj["E1"] = "total_blocking_time"
        sheet_obj["F1"] = "cumulative_layout_shift"
        sheet_obj["G1"] = "speed_index"
        sheet_obj["H1"] = "performance"

        # Установка ширины колонки
        # sheet_obj.column_dimensions['A'].width = 30
        sheet_obj.column_dimensions['B'].width = 90
        sheet_obj.column_dimensions['C'].width = 25
        sheet_obj.column_dimensions['D'].width = 25
        sheet_obj.column_dimensions['E'].width = 25
        sheet_obj.column_dimensions['F'].width = 25
        sheet_obj.column_dimensions['G'].width = 15
        sheet_obj.column_dimensions['H'].width = 15


    # Откріваем файл, если его нет создаем новый
    def open_or_create_xl(self, name):
        try:
            # Загрузка файла Excel
            work_book = load_workbook(name)
            print("open document")
            return work_book

        except FileNotFoundError:
            work_book = Workbook()
            print("create document")
            return work_book

    # Создание нового листа
    def add_new_sheet(self, sheet_name):
        new_sheet = self.work_book.create_sheet(sheet_name)
        # Назначаем новый лист активным
        self.work_book.active = new_sheet
        return new_sheet

    # Создание объекта для заполнения (цвет фона)
    def color_fill_call(self, color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')  # Например, желтый цвет

    # Запись данных в файл xlsx
    def save_xlsx(self, save_name):
        self.work_book.save(self.get_file(save_name))
        print('save file')


class CreateTable(Document):
    def __init__(self, open_file):
        super().__init__(open_file)
        self.num = 2 # записываем данные с 2-й строки

        # цвета заливки ячеек
        self.color_bad = 'ff3333'
        self.color_normal = 'ffaa33'
        self.color_good = '00cc66'

        self.sum_value_first_contentful_paint = 0
        self.sum_value_largest_contentful_paint = 0
        self.sum_value_total_blocking_time = 0
        self.sum_value_cumulative_layout_shift = 0
        self.sum_value_speed_index = 0
        self.sum_value_performance = 0

    first_contentful_paint = {
        "good": {"value": 1, "char": "<"}, "normal": [1, 2], "bad": {"value": 2, "char": ">"},
    }

    largest_contentful_paint = {
        "good": {"value": 2, "char": "<"}, "normal": [2, 4], "bad": {"value": 4, "char": ">"},
    }

    total_blocking_time = {
        "good": {"value": 100, "char": "<"}, "normal": [100, 300], "bad": {"value": 300, "char": ">"},
    }

    cumulative_layout_shift = {
        "good": {"value": 0.1, "char": "<"}, "normal": [0.1, 0.25], "bad": {"value": 0.25, "char": ">"},
    }

    speed_index = {
        "good": {"value": 1, "char": "<"}, "normal": [1, 1.5], "bad": {"value": 1.5, "char": ">"},
    }

    performance = {
        "good": {"value": 90, "char": ">="}, "normal": [50, 90], "bad": {"value": 50, "char": "<"},
    }

    # проверяем значение и возвращаем цвет ячейки
    def paint_color_call(self, data, value):
        # GOOD ------------------------------------------
        if data["good"]["char"] == ">=":
            if value >= data["good"]["value"]:
                return self.color_fill_call(self.color_good)

        if data["good"]["char"] == "<":
            if value < data["good"]["value"]:
                return self.color_fill_call(self.color_good)

        # HORMAL ----------------------------------------
        if value >= data["normal"][0] and value <= data["normal"][1]:
            return self.color_fill_call(self.color_normal)

        # BAD -------------------------------------------
        if data["bad"]["char"] == ">":
            if value > data["bad"]["value"]:
                return self.color_fill_call(self.color_bad)

        if data["bad"]["char"] == "<":
            if value < data["bad"]["value"]:
                return self.color_fill_call(self.color_bad)

    def complete_table(self, sheet_obj, file_name):
        open_file = os.path.join(self.path_save, f"{file_name}.json")
        with open(open_file, 'r') as json_file:
            file_json = json.load(json_file)

            for card_info in file_json:
                sheet_obj[f"A{self.num}"] = card_info["id"]
                sheet_obj[f"B{self.num}"] = card_info["url"]
                
                c_val = card_info["first_contentful_paint"]
                c_number = float(c_val.replace(' s', ''))
                sheet_obj[f"C{self.num}"] = c_val
                sheet_obj[f"C{self.num}"].fill = self.paint_color_call(self.first_contentful_paint, c_number)
                self.sum_value_first_contentful_paint += c_number

                d_val = card_info["largest_contentful_paint"]
                d_number = float(d_val.replace(' s', ''))
                sheet_obj[f"D{self.num}"] = d_val
                sheet_obj[f"D{self.num}"].fill = self.paint_color_call(self.largest_contentful_paint, d_number)
                self.sum_value_largest_contentful_paint += d_number
                
                e_val = card_info["total_blocking_time"]
                e_number = int(e_val.replace(' ms', '').replace(',', ''))
                sheet_obj[f"E{self.num}"] = str(e_number) + ' ms'
                sheet_obj[f"E{self.num}"].fill = self.paint_color_call(self.total_blocking_time ,int(e_number))
                self.sum_value_total_blocking_time += int(e_number)
                
                f_val = card_info["cumulative_layout_shift"]
                f_number = float(f_val)
                sheet_obj[f"F{self.num}"] = f_val
                sheet_obj[f"F{self.num}"].fill = self.paint_color_call(self.cumulative_layout_shift, f_number)
                self.sum_value_cumulative_layout_shift += f_number
                
                g_val = card_info["speed_index"]
                g_number = float(g_val.replace(' s', ''))
                sheet_obj[f"G{self.num}"] = g_val
                sheet_obj[f"G{self.num}"].fill = self.paint_color_call(self.speed_index, g_number)
                self.sum_value_speed_index += g_number

                h_val = card_info["performance"]
                h_number = int(h_val * 100)
                sheet_obj[f"H{self.num}"] = str(h_number)
                sheet_obj[f"H{self.num}"].fill = self.paint_color_call(self.performance, h_number)
                self.sum_value_performance += h_number

                self.num += 1

    # получаем средние значение и вносим в таблицу последнуй строкой
    def complete_average_value(self, sheet_obj):
        # получаем средние значение
        element_count = self.num - 2
        number_sum = self.num + 2

        c_average = self.sum_value_first_contentful_paint / element_count
        sheet_obj[f"C{number_sum}"] = str(round(c_average, 1)).replace(',', '.') + ' s'
        sheet_obj[f"C{number_sum}"].fill = self.paint_color_call(self.first_contentful_paint, c_average)

        d_average = self.sum_value_largest_contentful_paint / element_count
        sheet_obj[f"D{number_sum}"] = str(round(d_average, 1)).replace(',', '.') + ' s'
        sheet_obj[f"D{number_sum}"].fill = self.paint_color_call(self.largest_contentful_paint, d_average)

        e_average = self.sum_value_total_blocking_time / element_count
        sheet_obj[f"E{number_sum}"] = str(int(e_average)).replace(',', '.') + ' ms'
        sheet_obj[f"E{number_sum}"].fill = self.paint_color_call(self.total_blocking_time, e_average)

        f_average = self.sum_value_cumulative_layout_shift / element_count
        sheet_obj[f"F{number_sum}"] = str(round(f_average, 3)).replace(',', '.')
        sheet_obj[f"F{number_sum}"].fill = self.paint_color_call(self.cumulative_layout_shift, f_average)

        g_average = self.sum_value_speed_index / element_count
        sheet_obj[f"G{number_sum}"] = str(round(g_average, 1)).replace(',', '.') + ' s'
        sheet_obj[f"G{number_sum}"].fill = self.paint_color_call(self.speed_index, g_average)
        
        h_average = int(self.sum_value_performance / element_count)
        sheet_obj[f"H{number_sum}"] = str(h_average)
        sheet_obj[f"H{number_sum}"].fill = self.paint_color_call(self.performance, h_average)
