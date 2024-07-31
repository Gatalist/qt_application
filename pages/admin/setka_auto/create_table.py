from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd



class Document:
    def __init__(self, open_file):
        self.open_file = open_file
        self.work_book = self.open_or_create_xl(self.open_file)
        self.sheet = self.work_book.active

        self.data = pd.read_excel(self.open_file)
        self.column_name_list = self.data.columns.tolist()

        # цвета заливки ячеек
        self.color_basic = 'd9d2e9' # фиолетовый
        self.color_value = 'fff2cc' # желтый

        # Создание списка значений для выбора
        self.choices_type = ['-', 'Список', 'Множественное', 'Строка']
        self.choices_price = ['-', 'Да', 'Нет']
        self.choices_filter = ['-', 'Да', 'Нет']

    # добавляем первую строку, заголовки колонок
    def add_title_column(self, sheet_obj):
        # устанавливаем название колонок
        sheet_obj["A1"] = "№ группы"
        # цвета заливки ячеек
        sheet_obj["A1"].fill = self.color_fill_call(self.color_basic)
        # Центрирование текста в ячейке
        sheet_obj["A1"].alignment = Alignment(horizontal='center', vertical='center') 

        sheet_obj["B1"] = "Группа характеристик"
        sheet_obj["B1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["B1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["C1"] = "№ хар-ки"
        sheet_obj["C1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["C1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["D1"] = "Характеристика"
        sheet_obj["D1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["D1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["E1"] = "Значение (примеры)"
        sheet_obj["E1"].fill = self.color_fill_call(self.color_value)
        sheet_obj["E1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["F1"] = "Тип характеристики"
        sheet_obj["F1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["F1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["G1"] = "Ценники ( иконки в категории)"
        sheet_obj["G1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["G1"].alignment = Alignment(horizontal='center', vertical='center')

        sheet_obj["H1"] = "Отображение Фильтра в категории"
        sheet_obj["H1"].fill = self.color_fill_call(self.color_basic)
        sheet_obj["H1"].alignment = Alignment(horizontal='center', vertical='center')

        # Установка высоты первой строки
        sheet_obj.row_dimensions[1].height = 30

        # Установка ширины колонки
        sheet_obj.column_dimensions['A'].width = 10
        sheet_obj.column_dimensions['B'].width = 40
        sheet_obj.column_dimensions['C'].width = 15
        sheet_obj.column_dimensions['D'].width = 35
        sheet_obj.column_dimensions['E'].width = 45
        sheet_obj.column_dimensions['F'].width = 25
        sheet_obj.column_dimensions['G'].width = 35
        sheet_obj.column_dimensions['H'].width = 35

    def add_choices_to_call(self, sheet_obj, call, list_name):
        """ Добавляем выпадающий список значений в ячейку """
        # Создание объекта DataValidation
        data_validation = DataValidation(
            type="list", 
            formula1='"' + ','.join(list_name) + '"', 
            allow_blank=True)

        data_validation.add(sheet_obj[call])
        sheet_obj.add_data_validation(data_validation)
        # Первый элемент списка выбираем по умолчанию
        sheet_obj[call].value = list_name[0]

        return sheet_obj, data_validation


    # Открываем файл, если его нет создаем новый
    def open_or_create_xl(self, name):
        try:
            # Загрузка файла Excel
            work_book = load_workbook(name)
            print("open document")
            return work_book

        except FileNotFoundError:
            work_book = Workbook()
            print("create document")
            return work_book

    # Создание нового листа
    def add_new_sheet(self, sheet_name):
        new_sheet = self.work_book.create_sheet(sheet_name)
        # Назначаем новый лист активным
        self.work_book.active = new_sheet
        return new_sheet

    # Создание объекта для заполнения (цвет фона)
    def color_fill_call(self, color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')  # Например, желтый цвет

    # Запись данных в файл xlsx
    def save_xlsx(self, save_name):
        self.work_book.save(save_name)
        print('save file')


class CreateTable(Document):
    def __init__(self, open_file):
        super().__init__(open_file)
        self.num = 3

    def dell_id_name_from_list(self):
        """ Удаляем ID, Name из списка названий колонок """
        self.column_name_list.pop(0)
        self.column_name_list.pop(0)

    def get_data_column(self, column_name):
        """Получаем первые 3 уникальные значения с столбца (атрибута)"""
        data_column = self.data[column_name]
        data_col_3_row = list(set(data_column))[:3]
        unique_data_col = [str(elem).split(';')[0] for elem in data_col_3_row if str(elem) != "nan"]
        print(unique_data_col)
        return unique_data_col

    def strip_name_column(self, column_name: str):
        """Получаем название столбца (атрибута)"""
        return column_name.split('-')[0]
       

    def complete_table(self, sheet_obj, col_name):
        num_local = self.num
        
        call_d = f"D{num_local}"
        sheet_obj[call_d] = self.strip_name_column(col_name)
        sheet_obj[call_d].alignment = Alignment(horizontal='center', vertical='center')
        
        # Тип характеристики
        # Добавляем список выбора поля
        call_f = f"F{num_local}"
        self.add_choices_to_call(sheet_obj, call_f, self.choices_type)
        sheet_obj[call_f].alignment = Alignment(horizontal='center', vertical='center')

        # Ценники ( иконки в категории)
        # Добавляем список выбора поля
        call_f = f"G{num_local}"
        self.add_choices_to_call(sheet_obj, call_f, self.choices_price)
        sheet_obj[call_f].alignment = Alignment(horizontal='center', vertical='center')

        # Отображение Фильтра в категории
        # Добавляем список выбора поля
        call_f = f"H{num_local}"
        self.add_choices_to_call(sheet_obj, call_f, self.choices_price)
        sheet_obj[call_f].alignment = Alignment(horizontal='center', vertical='center')

        # Добавляем значение (примеры)
        for elem in self.get_data_column(col_name):
            sheet_obj[f"E{num_local}"] = elem
            sheet_obj[f"E{num_local}"].alignment = Alignment(horizontal='center', vertical='center')
            num_local += 1
        self.num += 4

    def for_to_colunm(self, sheet_obj):
        for col_name in self.column_name_list:
            self.complete_table(sheet_obj, col_name)


# # ------------------------ Записываем в xlsx документ
# pyxl = CreateTable(open_file='pro')

# # создаем лист на который будем записывать данные
# sheet_obj = pyxl.add_new_sheet('Гаджеты виртуальной реальности')

# # добавляем титульную строку
# pyxl.add_title_column(sheet_obj)

# # Удаляем ID, Name из списка названий колонок
# pyxl.dell_id_name_from_list()

# # добавляем результаты
# pyxl.for_to_colunm(sheet_obj)

# # сохраняем файл
# pyxl.save_xlsx("new")
